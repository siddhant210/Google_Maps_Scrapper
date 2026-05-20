from flask import Flask, render_template, request, send_file, jsonify
from scraper import scrape_google_maps
from whatsapp_service import (load_template, save_template, send_bulk_whatsapp,
                               sanitize_phone, save_attachment, clear_attachment, get_attachment_path)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, os

app = Flask(__name__)
scraped_data = []
os.makedirs("exports", exist_ok=True)


@app.route("/")
def home():
    return render_template("index.html", data=scraped_data, error=None)


@app.route("/scrape", methods=["GET", "POST"])
def scrape():
    global scraped_data
    if request.method == "GET":
        return render_template("index.html", data=scraped_data, error=None)
    query = request.form.get("query", "").strip()
    if not query:
        return render_template("index.html", data=[], error="Please enter a search query.")
    try:
        scraped_data = scrape_google_maps(query)
        if not scraped_data:
            return render_template("index.html", data=[],
                error="No results found. Try a different query.")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return render_template("index.html", data=[], error=f"Scraper error: {str(e)}")
    return render_template("index.html", data=scraped_data, error=None)


@app.route("/export")
def export():
    global scraped_data
    if not scraped_data:
        return "No data to export.", 400
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Leads"
    columns = ["Name","Category","Rating","Reviews","Latest Review","Email","Address","Phone","Website"]
    hf = PatternFill("solid", fgColor="1A1A2E")
    hfont = Font(name="Calibri", bold=True, color="00E5A0", size=11)
    ha = Alignment(horizontal="center", vertical="center")
    af = PatternFill("solid", fgColor="F7F9FC")
    nf = PatternFill("solid", fgColor="FFFFFF")
    cf = Font(name="Calibri", size=10, color="1A1A2E")
    ca = Alignment(vertical="center", wrap_text=True)
    thin = Side(style="thin", color="E0E4EC")
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 32
    for ci, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font=hfont; c.fill=hf; c.alignment=ha; c.border=b
    for ri, item in enumerate(scraped_data, 2):
        ws.row_dimensions[ri].height = 22
        fill = af if ri%2==0 else nf
        for ci, col in enumerate(columns, 1):
            val = item.get(col,"") or ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font=cf; cell.fill=fill; cell.alignment=ca; cell.border=b
            if col=="Website" and val.startswith("http"):
                cell.hyperlink=val
                cell.font=Font(name="Calibri",size=10,color="0563C1",underline="single")
            if col in ("Rating","Reviews"):
                cell.alignment=Alignment(horizontal="center",vertical="center")
    widths={"Name":28,"Category":18,"Rating":8,"Reviews":10,"Latest Review":22,"Email":25,"Address":45,"Phone":16,"Website":35}
    for ci,col in enumerate(columns,1):
        ws.column_dimensions[get_column_letter(ci)].width=widths[col]
    ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:{get_column_letter(len(columns))}1"
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="leads.xlsx")


# ── WhatsApp ──────────────────────────────────────────────────────────────────

@app.route("/whatsapp-template", methods=["GET","POST"])
def whatsapp_template():
    if request.method=="GET":
        return jsonify({"template": load_template()})
    t = (request.json or {}).get("template","").strip()
    if not t: return jsonify({"error":"Empty"}), 400
    return jsonify({"success":True}) if save_template(t) else (jsonify({"error":"Save failed"}),500)


@app.route("/whatsapp-lead-count")
def whatsapp_lead_count():
    global scraped_data
    total      = len(scraped_data)
    with_phone = sum(1 for l in scraped_data if sanitize_phone(l.get("Phone","")))
    return jsonify({"total": total, "with_phone": with_phone})


@app.route("/whatsapp-attachment", methods=["POST","DELETE"])
def whatsapp_attachment():
    if request.method=="DELETE":
        clear_attachment()
        return jsonify({"success":True})
    if "file" not in request.files:
        return jsonify({"error":"No file"}), 400
    file = request.files["file"]
    file.seek(0,2); size=file.tell(); file.seek(0)
    if size > 20*1024*1024:
        return jsonify({"error":"File too large (max 20MB)"}), 400
    path = save_attachment(file.read(), file.filename)
    return jsonify({"success":True, "filename":file.filename, "size":size, "path":path})


@app.route("/whatsapp-attachment-info")
def whatsapp_attachment_info():
    path = get_attachment_path()
    if path and os.path.exists(path):
        return jsonify({"has_attachment":True,"filename":os.path.basename(path),"size":os.path.getsize(path)})
    return jsonify({"has_attachment":False})


@app.route("/send-whatsapp", methods=["POST"])
def send_whatsapp():
    global scraped_data
    body        = request.json or {}
    delay       = int(body.get("delay", 5))
    test_number = body.get("test_number","").strip() or None
    chunk_start = body.get("chunk_start")   # 1-based, inclusive
    chunk_end   = body.get("chunk_end")     # 1-based, inclusive

    if not test_number and not scraped_data:
        return jsonify({"error":"No leads. Scrape first."}), 400

    # Convert to int if present
    if chunk_start: chunk_start = int(chunk_start)
    if chunk_end:   chunk_end   = int(chunk_end)

    try:
        results = send_bulk_whatsapp(
            scraped_data, load_template(),
            delay_between=delay,
            test_number=test_number,
            attachment_path=get_attachment_path(),
            chunk_start=chunk_start,
            chunk_end=chunk_end,
        )
        return jsonify(results)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True)
        